# cnt = count
# comp = component
# curr = current
# idx = index
# lang = language
# props = properties
# ui = user interface


from PyQt5.QtWidgets import (QMainWindow, QWidget, QLabel, QMenuBar, QMenu, QAction, QStatusBar,
                             QTableWidget, QTableWidgetItem, QPushButton, QAction, QSizePolicy,
                             QFileDialog, QAbstractItemView, QMessageBox, QListWidgetItem,
                             QHBoxLayout, QVBoxLayout, QHeaderView)
from PyQt5.QtCore import Qt, QCoreApplication
from PyQt5.QtGui import QFont, QIcon
from openpyxl import load_workbook
import traceback

from substance import Substance
from stream import Stream
from stream_properties_window import StreamProperties

conds_names = ["Temperature [C]", "Pressure [Pa]", "Mass Flow [kg/sec]", "Phase"]


class ComponentsLibrary(QMainWindow):
    def __init__(self, stream_name, worksheet_window):
        super().__init__()

        self.worksheet_window = worksheet_window
        self.streams_list = self.worksheet_window.ui.streams_list
        self.streams = self.worksheet_window.streams
        self.windows = self.worksheet_window.windows
        self.stream_name = stream_name

        self.comp_set = set()
        self.chosen_comps = {}

        self.ui = ComponentsLibraryUI()
        self.ui.setup_ui(self)

    def test_program_func(self):
        self.open_library_file()

        if self.worksheet_window.curr_lang == "English":
            self.chosen_comps["HC003"] = self.all_comps["HC003"]
        else:
            self.chosen_comps["УВ003"] = self.all_comps["УВ003"]

        stream = Stream(self.stream_name, conds_names, self.chosen_comps)
        # saving information about the stream in a dictionary for further use
        self.streams_list.addItem(QListWidgetItem(self.stream_name))
        self.streams[self.stream_name] = stream

        self.streams[self.stream_name].conds["Temperature [C]"] = "45"
        self.streams[self.stream_name].conds["Pressure [Pa]"] = "200000"
        self.streams[self.stream_name].conds["Mass Flow [kg/sec]"] = "7"
        self.streams[self.stream_name].conds["Phase"] = "empty (phase)"
        self.streams[self.stream_name].fracs["component 1"]["Mass Fraction"] = "1.0"

        if self.stream_name not in self.windows.keys():  # checks if a StreamProperties Window has been created for this stream
            self.windows[self.stream_name] = StreamProperties(self.stream_name, self.worksheet_window)

        if self.worksheet_window.curr_lang == "English":
            del self.chosen_comps["HC003"]
            self.stream_name = f"Stream {self.worksheet_window.ui.streams_list.count() + 1}"
            self.chosen_comps["HC007"] = self.all_comps["HC007"]
        else:
            del self.chosen_comps["УВ003"]
            self.stream_name = f"Поток {self.worksheet_window.ui.streams_list.count() + 1}"
            self.chosen_comps["УВ007"] = self.all_comps["УВ007"]

        stream = Stream(self.stream_name, conds_names, self.chosen_comps)
        # saving information about the stream in a dictionary for further use
        self.streams_list.addItem(QListWidgetItem(self.stream_name))
        self.streams[self.stream_name] = stream

        self.streams[self.stream_name].conds["Temperature [C]"] = "80"
        self.streams[self.stream_name].conds["Pressure [Pa]"] = "200000"
        self.streams[self.stream_name].conds["Mass Flow [kg/sec]"] = "3.31"
        self.streams[self.stream_name].conds["Phase"] = "empty (phase)"
        self.streams[self.stream_name].fracs["component 1"]["Mass Fraction"] = "1.0"

        if self.worksheet_window.curr_lang == "English":
            del self.chosen_comps["HC007"]
        else:
            del self.chosen_comps["УВ007"]

        if self.stream_name not in self.windows.keys():  # checks if a StreamProperties Window has been created for this stream
            self.windows[self.stream_name] = StreamProperties(self.stream_name, self.worksheet_window)

        self.close()
        self.deleteLater()

    def test_program_sep_func(self):
        self.open_library_file()

        if self.worksheet_window.curr_lang == "English":
            self.chosen_comps["HC001"] = self.all_comps["HC001"]
            self.chosen_comps["HC002"] = self.all_comps["HC002"]
            self.chosen_comps["HC003"] = self.all_comps["HC003"]
            self.chosen_comps["HC005"] = self.all_comps["HC005"]
            self.chosen_comps["HC007"] = self.all_comps["HC007"]
        else:
            self.chosen_comps["УВ001"] = self.all_comps["УВ001"]
            self.chosen_comps["УВ002"] = self.all_comps["УВ002"]
            self.chosen_comps["УВ003"] = self.all_comps["УВ003"]
            self.chosen_comps["УВ005"] = self.all_comps["УВ005"]
            self.chosen_comps["УВ007"] = self.all_comps["УВ007"]

        stream = Stream(self.stream_name, conds_names, self.chosen_comps)
        # saving information about the stream in a dictionary for further use
        self.streams_list.addItem(QListWidgetItem(self.stream_name))
        self.streams[self.stream_name] = stream

        self.streams[self.stream_name].conds["Temperature [C]"] = "20"
        self.streams[self.stream_name].conds["Pressure [Pa]"] = "100000"
        self.streams[self.stream_name].conds["Mass Flow [kg/sec]"] = "1"
        self.streams[self.stream_name].conds["Phase"] = "empty (phase)"
        self.streams[self.stream_name].fracs["component 1"]["Molar Fraction"] = "0.605"
        self.streams[self.stream_name].fracs["component 2"]["Molar Fraction"] = "0.175"
        self.streams[self.stream_name].fracs["component 3"]["Molar Fraction"] = "0.147"
        self.streams[self.stream_name].fracs["component 4"]["Molar Fraction"] = "0.044"
        self.streams[self.stream_name].fracs["component 5"]["Molar Fraction"] = "0.029"

        if self.worksheet_window.curr_lang == "English":
            del self.chosen_comps["HC001"]
            del self.chosen_comps["HC002"]
            del self.chosen_comps["HC003"]
            del self.chosen_comps["HC005"]
            del self.chosen_comps["HC007"]
        else:
            del self.chosen_comps["УВ001"]
            del self.chosen_comps["УВ002"]
            del self.chosen_comps["УВ003"]
            del self.chosen_comps["УВ005"]
            del self.chosen_comps["УВ007"]

        if self.stream_name not in self.windows.keys():  # checks if a StreamProperties Window has been created for this stream
            self.windows[self.stream_name] = StreamProperties(self.stream_name, self.worksheet_window)

        self.close()
        self.deleteLater()

    def test_program_liq_func(self):
        self.open_library_file()

        if self.worksheet_window.curr_lang == "English":
            self.chosen_comps["HC010"] = self.all_comps["HC010"]
            self.chosen_comps["HC011"] = self.all_comps["HC011"]
        else:
            self.chosen_comps["УВ010"] = self.all_comps["УВ010"]
            self.chosen_comps["УВ011"] = self.all_comps["УВ011"]

        stream = Stream(self.stream_name, conds_names, self.chosen_comps)
        # saving information about the stream in a dictionary for further use
        self.streams_list.addItem(QListWidgetItem(self.stream_name))
        self.streams[self.stream_name] = stream

        self.streams[self.stream_name].conds["Temperature [C]"] = "80"
        self.streams[self.stream_name].conds["Pressure [Pa]"] = "101325"
        self.streams[self.stream_name].conds["Mass Flow [kg/sec]"] = "12"
        self.streams[self.stream_name].conds["Phase"] = "empty (phase)"
        self.streams[self.stream_name].fracs["component 1"]["Molar Fraction"] = "0.5"
        self.streams[self.stream_name].fracs["component 2"]["Molar Fraction"] = "0.5"

        if self.stream_name not in self.windows.keys():  # checks if a StreamProperties Window has been created for this stream
            self.windows[self.stream_name] = StreamProperties(self.stream_name, self.worksheet_window)

        if self.worksheet_window.curr_lang == "English":
            del self.chosen_comps["HC010"]
            del self.chosen_comps["HC011"]
            self.stream_name = f"Stream {self.worksheet_window.ui.streams_list.count() + 1}"
            self.chosen_comps["HC010"] = self.all_comps["HC010"]
            self.chosen_comps["HC011"] = self.all_comps["HC011"]
        else:
            del self.chosen_comps["УВ010"]
            del self.chosen_comps["УВ011"]
            self.stream_name = f"Поток {self.worksheet_window.ui.streams_list.count() + 1}"
            self.chosen_comps["УВ010"] = self.all_comps["УВ010"]
            self.chosen_comps["УВ011"] = self.all_comps["УВ011"]

        stream = Stream(self.stream_name, conds_names, self.chosen_comps)
        # saving information about the stream in a dictionary for further use
        self.streams_list.addItem(QListWidgetItem(self.stream_name))
        self.streams[self.stream_name] = stream

        self.streams[self.stream_name].conds["Temperature [C]"] = "45"
        self.streams[self.stream_name].conds["Pressure [Pa]"] = "101325"
        self.streams[self.stream_name].conds["Mass Flow [kg/sec]"] = "7"
        self.streams[self.stream_name].conds["Phase"] = "empty (phase)"
        self.streams[self.stream_name].fracs["component 1"]["Molar Fraction"] = "0.1"
        self.streams[self.stream_name].fracs["component 2"]["Molar Fraction"] = "0.9"

        if self.worksheet_window.curr_lang == "English":
            del self.chosen_comps["HC010"]
            del self.chosen_comps["HC011"]
        else:
            del self.chosen_comps["УВ010"]
            del self.chosen_comps["УВ011"]

        if self.stream_name not in self.windows.keys():  # checks if a StreamProperties Window has been created for this stream
            self.windows[self.stream_name] = StreamProperties(self.stream_name, self.worksheet_window)

        self.close()
        self.deleteLater()

    def open_library_file(self):
        # standard_folder = "D:/Other prjects on Py/ChemTech/libs"
        # self.library_path = QFileDialog.getOpenFileName(self, "Open Library File", standard_folder)[0]

        if self.worksheet_window.curr_lang == "English":
            self.library_path = "./libs/Hydrocarbons_C1-C10_en.xlsx"
            self.headers = ["Substance Class", "ID", "Name", "Chemical Formula"]
        elif self.worksheet_window.curr_lang == "Russian":
            self.library_path = "./libs/Hydrocarbons_C1-C10_ru.xlsx"
            self.headers = ["Класс вещества", "ID", "Название", "Химическая формула"]
        self.load_data_library_table(self.library_path)

        translate = QCoreApplication.translate
        self.ui.statusbar.showMessage(translate("ComponentsLibrary Window", "Add components to stream"))
        self.ui.statusbar.setStyleSheet("QStatusBar{background:rgb(0,255,0);color:black;font-weight:bold;}")
        # when we loaded data, we understood what columns headers are needed, then we set headers for current stream table
        self.ui.curr_stream_table.setRowCount(0)
        self.ui.curr_stream_table.setColumnCount(self.cols_count)

        self.ui.curr_stream_table.setHorizontalHeaderLabels(self.headers)
        self.ui.curr_stream_table.resizeColumnsToContents()
        self.ui.curr_stream_table.setEditTriggers(QAbstractItemView.NoEditTriggers)

    def load_data_library_table(self, path):
        self.transfer_all_comps_to_dict(path)
        self.fill_library_table_with_values()

    def transfer_all_comps_to_dict(self, path):
        self.all_comps = {}

        workbook = load_workbook(path)
        sheets_names = workbook.sheetnames
        for sheet in sheets_names:
            sheet_vals = list(workbook[sheet].values)
            props_groups = {}
            for row_idx in range(len(sheet_vals)):
                if row_idx == 0:
                    props_cnt_in_props_group = 1
                    previous = sheet_vals[row_idx][0]
                    for val in sheet_vals[row_idx]:
                        if (val is not None) and (val not in props_groups.keys()):
                            props_groups[val] = props_cnt_in_props_group
                            props_cnt_in_props_group = 1
                            previous = val
                        else:
                            props_cnt_in_props_group += 1
                            props_groups[previous] = props_cnt_in_props_group
                elif row_idx == 1:
                    props = sheet_vals[row_idx]
                    counter = 0
                    for props_group, props_cnt_in_props_group in props_groups.items():
                        props_groups[props_group] = props[counter:counter + props_cnt_in_props_group]
                        counter += props_cnt_in_props_group
                else:
                    props_class = sheet
                    props_vals = sheet_vals[row_idx]
                    substance_class, substance_id = props_vals[0], props_vals[1]
                    comp_key = substance_class + str(substance_id)
                    if comp_key not in self.all_comps.keys():
                        self.all_comps[comp_key] = Substance()
                    counter = 0
                    for props_group, props in props_groups.items():
                        props_cnt = len(props)
                        self.all_comps[comp_key].get_props(props_class, props_group, props_vals[counter:counter + props_cnt])
                        counter += props_cnt

    def fill_library_table_with_values(self):
        self.rows_cnt = len(self.all_comps)  # components count in library file
        self.cols_count = len(self.headers)  # columns count is equal number of headers (component characteristics)
        self.ui.library_table.setRowCount(self.rows_cnt)
        self.ui.library_table.setColumnCount(self.cols_count)

        self.ui.library_table.setHorizontalHeaderLabels(self.headers)

        row_idx = 0
        for comp_key in self.all_comps.keys():
            col_idx = 0
            classification = self.all_comps[comp_key].general_info.classification
            classification_vals = list(classification.output_props().values())
            for val in classification_vals:
                val = QTableWidgetItem(str(val))
                val.setTextAlignment(Qt.AlignCenter)
                self.ui.library_table.setItem(row_idx, col_idx, QTableWidgetItem(val))
                col_idx += 1
            row_idx += 1

        self.ui.library_table.setEditTriggers(QAbstractItemView.NoEditTriggers)  # cells can not be changed now

    def add_comp_to_stream(self):
        curr_row = self.ui.library_table.currentRow()
        if curr_row != -1:  # if user don't choose the cell, current row is equal -1
            # comp_key = substance class + id
            comp_substance_class = self.ui.library_table.item(curr_row, 0).text()
            comp_id = str(self.ui.library_table.item(curr_row, 1).text())
            comp_key = comp_substance_class + comp_id
            if comp_key not in self.comp_set:  # if substance is already in the stream, don't add it again
                rows_cnt = self.ui.curr_stream_table.rowCount() + 1
                self.ui.curr_stream_table.setRowCount(rows_cnt)
                row_for_add = rows_cnt - 1
                for col_idx in range(0, self.ui.curr_stream_table.columnCount()):
                    chosen_comp_substance_class = self.ui.library_table.item(curr_row, 0).text()
                    chosen_comp_id = self.ui.library_table.item(curr_row, 1).text()
                    chosen_comp_key = chosen_comp_substance_class + chosen_comp_id
                    self.chosen_comps[chosen_comp_key] = self.all_comps[chosen_comp_key]
                    chosen_comp_classification = self.all_comps[chosen_comp_key].general_info.classification
                    classification_vals = list(chosen_comp_classification.output_props().values())
                    item = classification_vals[col_idx]
                    val = QTableWidgetItem(str(item))
                    val.setTextAlignment(Qt.AlignCenter)
                    self.ui.curr_stream_table.setItem(row_for_add, col_idx, QTableWidgetItem(val))
                self.comp_set.add(comp_key)

        self.ui.curr_stream_table.sortItems(1, Qt.AscendingOrder)
        self.ui.curr_stream_table.resizeColumnsToContents()

    def remove_comp_from_stream(self):
        curr_row = self.ui.curr_stream_table.currentRow()
        comp_key = str(self.ui.curr_stream_table.item(curr_row, 0).text()) + str(self.ui.curr_stream_table.item(curr_row, 1).text())

        del self.chosen_comps[comp_key]
        self.ui.curr_stream_table.removeRow(curr_row)
        self.comp_set.remove(comp_key)

    def add_stream_to_worksheet(self):
        # adding stream in streams list
        if self.ui.curr_stream_table.rowCount() == 0:  # check if stream is empty
            message_empty_curr_stream_table = QMessageBox(self)
            message_empty_curr_stream_table.setWindowTitle("Error")
            message_empty_curr_stream_table.setText("You haven't added any components")
            message_empty_curr_stream_table.show()
        else:
            stream = Stream(self.stream_name, conds_names, self.chosen_comps)
            # saving information about the stream in a dictionary for further use
            self.streams_list.addItem(QListWidgetItem(self.stream_name))
            self.streams[self.stream_name] = stream

            if self.stream_name not in self.windows.keys():  # checks if a StreamProperties Window has been created for this stream
                self.windows[self.stream_name] = StreamProperties(self.stream_name, self.worksheet_window)

            self.close()
            self.deleteLater()  # by my design, this line of code should delete the Components Library class object,
            # because I don't want to use RAM to store all information from every library I open


class ComponentsLibraryUI:
    def setup_ui(self, components_library):
        self.components_library = components_library

        self.components_library.resize(1600, 768)
        self.components_library.setWindowIcon(QIcon("./pics/program_icon.png"))
        # self.components_library.setMinimumSize(1600, 768)

        size_policy = QSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.central_widget = QWidget(self.components_library)
        self.central_widget.setSizePolicy(size_policy)
        self.components_library.setCentralWidget(self.central_widget)

        self.create_menubar(self.components_library)
        self.create_statusbar(self.components_library)
        self.create_widgets(self.components_library, size_policy)
        self.create_layout(self.central_widget)

        self.retranslate_ui(self.components_library)

    def create_menubar(self, parent):
        self.menubar = QMenuBar(parent)
        self.menubar.setFont(QFont("Segoe UI", 10))
        parent.setMenuBar(self.menubar)

        self.file_menu = QMenu(self.menubar)
        self.file_menu.setFont(QFont("Segoe UI", 10))
        self.action_open_file = QAction(parent)
        self.action_test = QAction(parent)
        self.action_test_sep = QAction(parent)
        self.action_test_liq = QAction(parent)

        self.file_menu.addAction(self.action_open_file)
        self.file_menu.addAction(self.action_test)
        self.file_menu.addAction(self.action_test_sep)
        self.file_menu.addAction(self.action_test_liq)

        self.action_open_file.triggered.connect(self.components_library.open_library_file)
        self.action_test.triggered.connect(self.components_library.test_program_func)
        self.action_test_sep.triggered.connect(self.components_library.test_program_sep_func)
        self.action_test_liq.triggered.connect(self.components_library.test_program_liq_func)

        self.menubar.addMenu(self.file_menu)

    def create_statusbar(self, parent):
        self.statusbar = QStatusBar(parent)
        self.statusbar.setStyleSheet("QStatusBar{background:rgb(255,0,0);color:black;font-weight:bold;}")
        parent.setStatusBar(self.statusbar)

    def create_widgets(self, parent, s_policy):
        self.library_table = QTableWidget(parent)
        self.library_table.setFont(QFont("Segoe UI", 10))
        h_header = self.library_table.horizontalHeader()
        h_header.setSectionResizeMode(QHeaderView.Stretch)
        v_header = self.library_table.verticalHeader()
        v_header.setSectionResizeMode(QHeaderView.Fixed)

        self.curr_stream_table = QTableWidget(parent)
        self.curr_stream_table.setFont(QFont("Segoe UI", 10))
        h_header = self.curr_stream_table.horizontalHeader()
        h_header.setSectionResizeMode(QHeaderView.Stretch)
        v_header = self.curr_stream_table.verticalHeader()
        v_header.setSectionResizeMode(QHeaderView.Fixed)

        self.library_table.doubleClicked.connect(self.components_library.add_comp_to_stream)
        self.curr_stream_table.doubleClicked.connect(self.components_library.remove_comp_from_stream)

        self.btn_add_comp_to_stream = QPushButton(parent)
        self.btn_add_comp_to_stream.setFont(QFont("Segoe UI", 10))
        self.btn_add_comp_to_stream.setFixedHeight(75)
        self.btn_add_comp_to_stream.setSizePolicy(s_policy)

        self.btn_remove_comp_from_stream = QPushButton(parent)
        self.btn_remove_comp_from_stream.setFont(QFont("Segoe UI", 10))
        self.btn_remove_comp_from_stream.setFixedHeight(75)
        self.btn_remove_comp_from_stream.setSizePolicy(s_policy)

        self.btn_add_stream = QPushButton(parent)
        self.btn_add_stream.setFont(QFont("Segoe UI", 10))
        self.btn_add_stream.setFixedHeight(50)
        self.btn_add_stream.setSizePolicy(s_policy)

        self.btn_add_comp_to_stream.clicked.connect(self.components_library.add_comp_to_stream)
        self.btn_remove_comp_from_stream.clicked.connect(self.components_library.remove_comp_from_stream)
        self.btn_add_stream.clicked.connect(self.components_library.add_stream_to_worksheet)

    def create_layout(self, parent):
        self.layout = QHBoxLayout(parent)

        btn_layuot = QVBoxLayout(parent)
        btn_layuot.addWidget(self.btn_add_comp_to_stream)
        btn_layuot.addWidget(self.btn_remove_comp_from_stream)

        right_layout = QVBoxLayout(parent)
        right_layout.addWidget(self.curr_stream_table)
        right_layout.addWidget(self.btn_add_stream)

        self.layout.addWidget(self.library_table)
        self.layout.addLayout(btn_layuot)
        self.layout.addLayout(right_layout)

        self.layout.setStretchFactor(self.library_table, 2)
        self.layout.setStretchFactor(btn_layuot, 1)
        self.layout.setStretchFactor(right_layout, 2)

        parent.setLayout(self.layout)

    def retranslate_ui(self, parent):
        translate = QCoreApplication.translate

        parent.setWindowTitle(translate("ComponentsLibrary Window", "Components Library"))
        self.file_menu.setTitle(translate("ComponentsLibrary Window", "File"))
        self.action_open_file.setText(translate("ComponentsLibrary Window", "Open"))
        self.action_open_file.setShortcut(translate("ComponentsLibrary Window", "Ctrl+Q"))
        self.action_test.setText(translate("ComponentsLibrary Window", "Test Vapour Streams"))
        self.action_test.setShortcut(translate("ComponentsLibrary Window", "Ctrl+1"))
        self.action_test_sep.setText(translate("ComponentsLibrary Window", "Test Sep"))
        self.action_test_sep.setShortcut(translate("ComponentsLibrary Window", "Ctrl+2"))
        self.action_test_liq.setText(translate("ComponentsLibrary Window", "Test Liquids Streams"))
        self.action_test_liq.setShortcut(translate("ComponentsLibrary Window", "Ctrl+3"))
        self.statusbar.showMessage(translate("ComponentsLibrary Window", "Open Library File"))
        self.btn_add_comp_to_stream.setText(translate("ComponentsLibrary Window", "Add"))
        self.btn_remove_comp_from_stream.setText(translate("ComponentsLibrary Window", "Remove"))
        self.btn_add_stream.setText(translate("ComponentsLibrary Window", "Add stream"))
